"""test_tracker_phase5.py — Wave 0 RED tests for CON-12/13 (split extract_job_id),
CON-14 (skipped_stale rename — local var only, dict key unchanged), CON-15/20
(user-column preservation in _write_tracker round-trip).

CRITICAL — Pitfall 6: The returned dict key from append_rows() is "flagged_stale"
and MUST stay "flagged_stale". Only the LOCAL variable `skipped_stale` is renamed
to `flagged_stale_count` in CON-14. Tests assert key == "flagged_stale".

Run with:
    ~/.job-scout-venv/bin/python3 -m pytest tests/test_tracker_phase5.py -x -q
"""

import json
import sys
from pathlib import Path

import openpyxl
import pytest

PROJECT_ROOT = Path(__file__).resolve().parent.parent
SCRIPTS_DIR = PROJECT_ROOT / "scripts"
sys.path.insert(0, str(SCRIPTS_DIR))


# ---------------------------------------------------------------------------
# Shared fixture: tracker with a user-added "My Notes" column
# ---------------------------------------------------------------------------

@pytest.fixture
def tracker_with_user_col(tmp_path):
    """Create a tracker xlsx with a user-added 'My Notes' column in column 17."""
    from tracker_utils import create_empty_tracker, HEADERS
    path = tmp_path / "JobScout_Tracker.xlsx"
    create_empty_tracker(str(path))
    wb = openpyxl.load_workbook(str(path))
    ws = wb.active
    ws.cell(row=1, column=len(HEADERS) + 1, value="My Notes")
    ws.cell(row=2, column=len(HEADERS) + 1, value="Important lead")
    wb.save(str(path))
    return path, len(HEADERS) + 1  # return path + the user-column index


# ---------------------------------------------------------------------------
# CON-13: extract_linkedin_job_id — LinkedIn-anchored extraction
# ---------------------------------------------------------------------------

def test_extract_linkedin_job_id_linkedin_url():
    """CON-13: extract_linkedin_job_id returns int job ID for valid LinkedIn view URL."""
    from tracker_utils import extract_linkedin_job_id  # RED until Plan 05-03 lands

    result = extract_linkedin_job_id("https://linkedin.com/jobs/view/3950000000001")
    assert result == 3950000000001, (
        f"Expected int 3950000000001, got: {result!r} (type: {type(result).__name__})"
    )
    assert isinstance(result, int), f"Expected int, got {type(result).__name__}"


def test_extract_linkedin_job_id_non_linkedin_returns_none():
    """CON-13: extract_linkedin_job_id returns None for non-LinkedIn URLs."""
    from tracker_utils import extract_linkedin_job_id  # RED until Plan 05-03 lands

    # Greenhouse URL — has no LinkedIn
    assert extract_linkedin_job_id("https://boards.greenhouse.io/airbnb/jobs/7890") is None, (
        "Greenhouse URL should return None"
    )
    # Lever URL — has no LinkedIn
    assert extract_linkedin_job_id("https://jobs.lever.co/stripe/abc-123") is None, (
        "Lever URL should return None"
    )
    # None input
    assert extract_linkedin_job_id(None) is None, (
        "None input should return None"
    )


def test_extract_linkedin_job_id_search_path():
    """CON-13: extract_linkedin_job_id supports linkedin.com/jobs/search/?currentJobId=... path."""
    from tracker_utils import extract_linkedin_job_id  # RED until Plan 05-03 lands

    result = extract_linkedin_job_id(
        "https://www.linkedin.com/jobs/search/?currentJobId=3950000000099"
    )
    assert result == 3950000000099, (
        f"Expected 3950000000099 from search path, got: {result!r}"
    )


# ---------------------------------------------------------------------------
# CON-13: extract_dedup_key — stable cross-source dedup key
# ---------------------------------------------------------------------------

def test_extract_dedup_key_linkedin_returns_id_string():
    """CON-13: extract_dedup_key returns str(linkedin_id) for LinkedIn URLs."""
    from tracker_utils import extract_dedup_key  # RED until Plan 05-03 lands

    result = extract_dedup_key("https://linkedin.com/jobs/view/3950000000001")
    assert result == "3950000000001", (
        f"Expected string '3950000000001', got: {result!r} (type: {type(result).__name__})"
    )
    assert isinstance(result, str), f"extract_dedup_key must return str, got {type(result).__name__}"


def test_extract_dedup_key_non_linkedin_returns_url():
    """CON-13: extract_dedup_key returns lowercased+stripped URL for non-LinkedIn URLs."""
    from tracker_utils import extract_dedup_key  # RED until Plan 05-03 lands

    url = "https://boards.greenhouse.io/airbnb/jobs/7890"
    result = extract_dedup_key(url)
    assert result == url, (
        f"Non-LinkedIn URL should be returned as-is (lowercased), got: {result!r}"
    )


def test_extract_dedup_key_none_returns_none():
    """CON-13: extract_dedup_key returns None for None or empty string inputs."""
    from tracker_utils import extract_dedup_key  # RED until Plan 05-03 lands

    assert extract_dedup_key(None) is None, "None input should return None"
    assert extract_dedup_key("") is None, "Empty string input should return None"


# ---------------------------------------------------------------------------
# CON-14 (Pitfall 6): dict key stays "flagged_stale", not "flagged_stale_count"
# ---------------------------------------------------------------------------

def test_flagged_stale_count_var(tmp_path):
    """CON-14 / Pitfall 6: dict key MUST be 'flagged_stale'; rename is local-var only."""
    from tracker_utils import create_empty_tracker, append_rows

    tracker_path = tmp_path / "JobScout_Tracker.xlsx"
    create_empty_tracker(str(tracker_path))
    new_rows = tmp_path / "new_rows.json"
    new_rows.write_text(json.dumps([{
        "date_found": "2026-04-28",
        "job_title": "VP Eng",
        "company": "Acme",
        "location": "Remote",
        "job_url": "https://linkedin.com/jobs/view/3950000099999",
        "tier": "A",
        "score": 80,
        "status": "New",
    }]))
    result = append_rows(str(tracker_path), str(new_rows))
    assert "flagged_stale" in result, "Pitfall 6 — dict key MUST be 'flagged_stale'"
    assert "flagged_stale_count" not in result, (
        "Pitfall 6 — local var rename must NOT leak to dict key"
    )
    assert "skipped_stale" not in result, (
        "Pitfall 6 — old key 'skipped_stale' must be gone (cleanup in CON-14)"
    )


# ---------------------------------------------------------------------------
# CON-15, CON-20: User-column preservation round-trip
# ---------------------------------------------------------------------------

def test_user_column_preservation_round_trip(tmp_path):
    """CON-15, CON-20: User-added column header + value survive an append_rows() round-trip.

    Protocol (from PLAN 05-01 spec):
    1. create_empty_tracker → row 1 = headers only
    2. append_rows with 1 scout row → row 2 = scout data
    3. openpyxl: set row 1 col 17 = "My Notes", row 2 col 17 = "Important lead"
    4. append_rows with another scout row → row 3 = new scout data
    5. Re-open: assert row 1 col 17 == "My Notes", row 2 col 17 == "Important lead"
    """
    from tracker_utils import create_empty_tracker, append_rows, HEADERS

    tracker_path = tmp_path / "JobScout_Tracker.xlsx"
    create_empty_tracker(str(tracker_path))

    # Step 2: pre-populate with 1 scout row
    prepop_json = tmp_path / "prepop.json"
    prepop_json.write_text(json.dumps([{
        "date_found": "2026-04-28",
        "job_title": "Software Engineer",
        "company": "Acme Corp",
        "location": "San Francisco, CA",
        "job_url": "https://linkedin.com/jobs/view/3950000000001",
        "tier": "B",
        "score": 75,
        "status": "New",
    }]))
    append_rows(str(tracker_path), str(prepop_json))

    # Step 3: add user column
    wb = openpyxl.load_workbook(str(tracker_path))
    ws = wb.active
    user_col = len(HEADERS) + 1
    ws.cell(row=1, column=user_col, value="My Notes")
    ws.cell(row=2, column=user_col, value="Important lead")
    wb.save(str(tracker_path))

    # Step 4: append another scout row
    new_json = tmp_path / "new_row.json"
    new_json.write_text(json.dumps([{
        "date_found": "2026-04-28",
        "job_title": "Product Manager",
        "company": "Example Inc",
        "location": "Remote",
        "job_url": "https://linkedin.com/jobs/view/3950000000002",
        "tier": "A",
        "score": 82,
        "status": "New",
    }]))
    append_rows(str(tracker_path), str(new_json))

    # Step 5: re-open and assert user column survived
    wb2 = openpyxl.load_workbook(str(tracker_path))
    ws2 = wb2.active
    assert ws2.cell(row=1, column=user_col).value == "My Notes", (
        f"CON-20: User column header 'My Notes' was lost after append_rows. "
        f"Got: {ws2.cell(row=1, column=user_col).value!r}"
    )
    assert ws2.cell(row=2, column=user_col).value == "Important lead", (
        f"CON-20: User column value 'Important lead' was lost after append_rows. "
        f"Got: {ws2.cell(row=2, column=user_col).value!r}"
    )


def test_user_column_preservation_after_multiple_appends(tmp_path):
    """CON-20: User-added column survives THREE consecutive append_rows() calls."""
    from tracker_utils import create_empty_tracker, append_rows, HEADERS

    tracker_path = tmp_path / "JobScout_Tracker.xlsx"
    create_empty_tracker(str(tracker_path))

    # Pre-populate with 1 row so we can set the user column value in row 2
    prepop_json = tmp_path / "prepop.json"
    prepop_json.write_text(json.dumps([{
        "date_found": "2026-04-28",
        "job_title": "Initial Role",
        "company": "Acme Corp",
        "location": "SF",
        "job_url": "https://linkedin.com/jobs/view/3950000000010",
        "tier": "B",
        "score": 72,
        "status": "New",
    }]))
    append_rows(str(tracker_path), str(prepop_json))

    # Set user column
    wb = openpyxl.load_workbook(str(tracker_path))
    ws = wb.active
    user_col = len(HEADERS) + 1
    ws.cell(row=1, column=user_col, value="My Notes")
    ws.cell(row=2, column=user_col, value="Important lead")
    wb.save(str(tracker_path))

    # Three more appends
    for i, (job_id, title) in enumerate([
        (3950000000011, "Role Alpha"),
        (3950000000012, "Role Beta"),
        (3950000000013, "Role Gamma"),
    ], start=1):
        batch_json = tmp_path / f"batch_{i}.json"
        batch_json.write_text(json.dumps([{
            "date_found": "2026-04-28",
            "job_title": title,
            "company": f"Company {i}",
            "location": "Remote",
            "job_url": f"https://linkedin.com/jobs/view/{job_id}",
            "tier": "C",
            "score": 60,
            "status": "New",
        }]))
        append_rows(str(tracker_path), str(batch_json))

    # After 3 appends, user column must still be intact
    wb2 = openpyxl.load_workbook(str(tracker_path))
    ws2 = wb2.active
    assert ws2.cell(row=1, column=user_col).value == "My Notes", (
        f"CON-20: User column header lost after 3 appends. "
        f"Got: {ws2.cell(row=1, column=user_col).value!r}"
    )
    assert ws2.cell(row=2, column=user_col).value == "Important lead", (
        f"CON-20: User column value lost after 3 appends. "
        f"Got: {ws2.cell(row=2, column=user_col).value!r}"
    )
