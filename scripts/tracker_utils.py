#!/usr/bin/env python3
"""
tracker_utils.py — Deterministic tracker management for Job Scout.

Handles reading, deduplicating, formatting, and writing the JobScout_Tracker.xlsx.
The SKILL.md calls this script instead of reimplementing formatting logic each run.
This ensures identical colors, deduplication, and structure every single time.

Usage:
    # Load existing tracker and get dedup set (returns JSON list of existing job IDs)
    python3 tracker_utils.py dedup-set /path/to/JobScout_Tracker.xlsx

    # Append new rows from a JSON file (deduplicates automatically)
    python3 tracker_utils.py append /path/to/JobScout_Tracker.xlsx /path/to/new_rows.json

    # Rebuild formatting on existing tracker (fix colors, remove dupes)
    python3 tracker_utils.py rebuild /path/to/JobScout_Tracker.xlsx
"""

import sys
import os
import json
import re
from datetime import datetime

try:
    import openpyxl
    from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print(
        "ERROR: openpyxl not installed. Install with: "
        "python3 -m venv ~/.job-scout-venv && source ~/.job-scout-venv/bin/activate && pip install openpyxl"
        "  (or: pip install --user openpyxl)",
        file=sys.stderr,
    )
    sys.exit(1)

# Allow running this script directly from the plugin root.
SCRIPTS_DIR = os.path.dirname(os.path.abspath(__file__))
if SCRIPTS_DIR not in sys.path:
    sys.path.insert(0, SCRIPTS_DIR)

# Column schema lives in schema.py — single source of truth.
from schema import (
    TRACKER_COLUMNS as HEADERS,
    TRACKER_COL_WIDTHS as COL_WIDTHS,
    TRACKER_JSON_KEYS,
    STALE_LINKEDIN_JOB_ID_THRESHOLD as STALE_JOB_ID_THRESHOLD,
    normalize_application_status,
)


# === FORMATTING CONSTANTS ===
# These NEVER change. Every run uses exactly these values.

# Exact hex colors — same every run, no variation
HEADER_FILL = PatternFill(start_color="2F5496", end_color="2F5496", fill_type="solid")
HEADER_FONT = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
A_TIER_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
B_TIER_FILL = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
C_TIER_FILL = PatternFill(start_color="F2DCDB", end_color="F2DCDB", fill_type="solid")  # Pink
STALE_FILL = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")   # Gray
NO_FILL = PatternFill()

THIN_BORDER = Border(
    left=Side(style='thin'), right=Side(style='thin'),
    top=Side(style='thin'), bottom=Side(style='thin')
)


def extract_linkedin_job_id(url):
    """Extract numeric LinkedIn job ID — anchored to linkedin.com URLs only.
    Returns int or None. Returns None for non-LinkedIn URLs (career-page,
    ATS board URLs, etc.) so stale-flag and dedup skip non-LinkedIn rows.

    CON-13: replaces the old generic extract_job_id which matched any 10+ digit
    run in any URL (false positives on ATS job IDs like greenhouse.io/jobs/7890).
    """
    if not url:
        return None
    s = str(url)
    # Handle /jobs/search/?currentJobId=<id> query-string form
    match = re.search(r'linkedin\.com/jobs/search/[^?]*\?(?:[^&]*&)*currentJobId=(\d{10,})', s)
    if match:
        return int(match.group(1))
    # Handle /jobs/view/<id> and /jobs/search/<non-digits><id> path forms
    match = re.search(r'linkedin\.com/jobs/(?:view|search)/\D*?(\d{10,})', s)
    return int(match.group(1)) if match else None


def extract_dedup_key(url):
    """Return a stable dedup key for any URL.

    For LinkedIn URLs, returns the numeric job ID as a string.
    For other URLs, returns the URL itself (lowercased, stripped).
    Used by rebuild() to deduplicate non-LinkedIn rows by full URL.

    CON-13: enables cross-source dedup — ATS URLs deduped by URL string,
    LinkedIn by job ID.
    """
    if not url:
        return None
    s = str(url).strip()
    if not s:
        return None
    linkedin_id = extract_linkedin_job_id(s)
    if linkedin_id is not None:
        return str(linkedin_id)
    return s.lower()


def extract_job_id(url):
    """DEPRECATED (CON-13): use extract_linkedin_job_id (LinkedIn-anchored) or
    extract_dedup_key (URL-as-string fallback). Kept for back-compat; removed in Phase 6.
    """
    return extract_linkedin_job_id(url)


def is_stale_by_id(url):
    """Check if a listing is likely stale based on LinkedIn job ID."""
    job_id = extract_linkedin_job_id(url)
    if job_id and job_id < STALE_JOB_ID_THRESHOLD:
        return True, job_id
    return False, job_id


def get_row_fill(tier, status, notes):
    """Determine the fill color for a row. Deterministic — same inputs always produce same output."""
    status_str = str(status or "")
    notes_str = str(notes or "")

    if "Stale" in status_str or "STALE" in notes_str:
        return STALE_FILL
    elif tier == "A":
        return A_TIER_FILL
    elif tier == "B":
        return B_TIER_FILL
    elif tier == "C":
        return C_TIER_FILL
    return NO_FILL


def create_empty_tracker(filepath):
    """Create a new empty tracker with proper headers and formatting."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Tracker"

    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}1"

    wb.save(filepath)
    return wb


def load_tracker(filepath):
    """Load existing tracker. Returns (workbook, list of rows, set of existing job IDs, user_extra_headers).

    user_extra_headers: list of column header names from row 1 beyond len(HEADERS).
    Empty list if no user-added columns exist.

    CON-20: 4-tuple return so _write_tracker can preserve user-added xlsx columns.
    """
    if not os.path.exists(filepath):
        wb = create_empty_tracker(filepath)
        return wb, [], set(), []

    wb = openpyxl.load_workbook(filepath)
    ws = wb.active

    # CON-20: discover user-added column headers from row 1 (beyond canonical HEADERS)
    ws_headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
    user_extra_headers = [h for h in ws_headers[len(HEADERS):] if h is not None]

    rows = []
    job_ids = set()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        row_list = list(row)
        if len(row_list) < len(HEADERS):
            row_list.extend([None] * (len(HEADERS) - len(row_list)))
        # CON-20: keep extra columns in row_list intact — _write_tracker will preserve them
        rows.append(row_list)

        url = row_list[9]  # Job URL column
        job_id = extract_linkedin_job_id(url)  # CON-13: LinkedIn-anchored only
        if job_id:
            job_ids.add(job_id)

    return wb, rows, job_ids, user_extra_headers


def get_dedup_set(filepath):
    """Return JSON list of existing job IDs for the SKILL.md to use for pre-search deduplication."""
    _, _, job_ids, _ = load_tracker(filepath)
    return sorted(list(job_ids))


def append_rows(filepath, new_rows_json_path):
    """
    Append new rows to the tracker. Deduplicates by job ID.
    new_rows.json format:
    [
        {
            "date_found": "2026-03-18",
            "job_title": "VP Engineering",
            "company": "Acme Corp",
            "location": "Seattle, WA",
            "comp_range": "$200K-$300K",
            "score": 78,
            "tier": "A",
            "connections": 3,
            "match_notes": "Strong commerce fit",
            "job_url": "https://linkedin.com/jobs/view/1234567890",
            "resume_tailored": "No",
            "resume_file": "",
            "status": "New",
            "notes": ""
        }
    ]
    """
    with open(new_rows_json_path, 'r') as f:
        new_rows = json.load(f)

    _, existing_rows, existing_ids, user_extra_headers = load_tracker(filepath)

    added = 0
    skipped_dupe = 0
    flagged_stale_count = 0  # CON-14: local var (dict key stays "flagged_stale" per Pitfall 6)

    for row_dict in new_rows:
        url = row_dict.get("job_url", "")
        job_id = extract_linkedin_job_id(url)  # CON-13: LinkedIn-anchored only

        # Dedup check
        if job_id and job_id in existing_ids:
            skipped_dupe += 1
            continue

        # Stale check
        stale, _ = is_stale_by_id(url)
        if stale:
            row_dict["status"] = "Stale — Verify"
            row_dict["notes"] = f"LIKELY STALE — LinkedIn job ID {job_id} suggests old listing. {row_dict.get('notes', '')}".strip()
            flagged_stale_count += 1

        # CON-02: validate application_status against STATUS_VALUES.
        # Warn-and-pass-through per locked decision — never rejects a row,
        # and never rewrites user data. Unknown values are preserved verbatim
        # in the row; a stderr warning fires so the user can fix the typo or
        # extend STATUS_VALUES if intentional.
        #
        # Default for absent `status` is "New" (preserves v=3 behavior). "New" is
        # a canonical member, so normalize_application_status returns (canonical, False)
        # and no warning fires. Case-insensitive matches (e.g., "DEAD" -> "Dead") are
        # silently canonicalized AND warned. Truly unknown strings ("Stale — Verify")
        # are preserved as-is AND warned.
        raw_status = row_dict.get("status", "New")
        canonical_status, status_coerced = normalize_application_status(raw_status)
        if status_coerced and raw_status:
            print(
                f"WARNING: row {added}: application_status {raw_status!r} not an exact "
                f"match in STATUS_VALUES; preserved as {canonical_status!r} "
                f"(add to STATUS_VALUES if intentional)",
                file=sys.stderr,
            )
        row_dict["status"] = canonical_status

        row_list = [
            row_dict.get("date_found", datetime.now().strftime("%Y-%m-%d")),
            row_dict.get("job_title", ""),
            row_dict.get("company", ""),
            row_dict.get("location", ""),
            row_dict.get("comp_range", ""),
            row_dict.get("score", 0),
            row_dict.get("tier", "C"),
            row_dict.get("connections", 0),
            row_dict.get("match_notes", ""),
            row_dict.get("job_url", ""),
            row_dict.get("resume_tailored", "No"),
            row_dict.get("resume_file", ""),
            row_dict.get("status", "New"),
            row_dict.get("notes", ""),
            row_dict.get("source", ""),         # v0.4 SCH-04: ats:greenhouse|...|linkedin
            row_dict.get("ats_provider", ""),   # v0.4 SCH-04: ats:greenhouse|... or empty
        ]

        existing_rows.append(row_list)
        if job_id:
            existing_ids.add(job_id)
        added += 1

    # Rebuild the entire file with consistent formatting
    _write_tracker(filepath, existing_rows, user_extra_headers=user_extra_headers)

    result = {
        "added": added,
        "skipped_duplicate": skipped_dupe,
        "flagged_stale": flagged_stale_count,  # CON-14: dict KEY stays "flagged_stale" (Pitfall 6)
        "total_rows": len(existing_rows)
    }
    return result


def rebuild(filepath):
    """Rebuild an existing tracker: remove dupes, fix colors, fix formatting."""
    _, rows, _, user_extra_headers = load_tracker(filepath)  # CON-20: unpack 4-tuple

    # Deduplicate — use extract_dedup_key (CON-13) so ATS/career-page rows dedup by URL string
    seen_keys = set()
    deduped = []
    dupes = 0

    for row in rows:
        url = row[9] if len(row) > 9 else None
        key = extract_dedup_key(url)  # CON-13: LinkedIn → job ID str; others → URL str

        if key and key in seen_keys:
            dupes += 1
            continue
        if key:
            seen_keys.add(key)
        deduped.append(row)

    _write_tracker(filepath, deduped, user_extra_headers=user_extra_headers)

    return {
        "original_rows": len(rows),
        "after_dedup": len(deduped),
        "duplicates_removed": dupes
    }


def _write_tracker(filepath, rows, user_extra_headers=None):
    """Write tracker with guaranteed consistent formatting.

    user_extra_headers: list of user-added column header names to re-emit in row 1
    at columns past len(HEADERS). For each data row, values at those indices are
    written through as plain cells (no scout formatting). CON-20 fix: replaces the
    old `break` that silently dropped user-added xlsx columns on every rewrite.
    """
    user_extra_headers = user_extra_headers or []
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Job Tracker"

    # Canonical headers (cols 1..len(HEADERS))
    for col, header in enumerate(HEADERS, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    # CON-20: re-emit user-added column headers in row 1 (cols past len(HEADERS))
    for i, extra_header in enumerate(user_extra_headers, start=len(HEADERS) + 1):
        cell = ws.cell(row=1, column=i, value=extra_header)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = THIN_BORDER

    # Column widths (canonical columns only — user columns get default width)
    for i, w in enumerate(COL_WIDTHS, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Data rows
    HYPERLINK_FONT = Font(color="0563C1", underline="single")  # Excel default hyperlink blue
    TIER_FONT = Font(bold=True)

    for r_idx, row in enumerate(rows, 2):
        tier = str(row[6]).strip() if len(row) > 6 and row[6] else ""
        status = str(row[12]).strip() if len(row) > 12 and row[12] else ""
        notes = str(row[13]).strip() if len(row) > 13 and row[13] else ""
        row_fill = get_row_fill(tier, status, notes)

        for col, val in enumerate(row, 1):
            cell = ws.cell(row=r_idx, column=col, value=val)
            cell.fill = row_fill
            cell.border = THIN_BORDER
            cell.alignment = Alignment(wrap_text=True, vertical='top')

            if col > len(HEADERS):
                # CON-20: user-added column — plain write-through, no scout formatting
                continue

            if col == 6:  # Score — center
                cell.alignment = Alignment(horizontal='center', vertical='top')
            elif col == 7:  # Tier — center + bold
                cell.alignment = Alignment(horizontal='center', vertical='top')
                cell.font = TIER_FONT
            elif col == 10 and val and isinstance(val, str) and val.startswith(('http://', 'https://')):
                # Job URL — make clickable hyperlink (blue + underline, native click)
                cell.hyperlink = val
                cell.font = HYPERLINK_FONT

    # Freeze + filter
    ws.freeze_panes = "A2"
    last_col = get_column_letter(len(HEADERS))
    ws.auto_filter.ref = f"A1:{last_col}{len(rows) + 1}"

    wb.save(filepath)


# === CLI ===

if __name__ == "__main__":
    if len(sys.argv) < 3:
        print("Usage: python3 tracker_utils.py <command> <tracker_path> [args...]")
        print("Commands: dedup-set, append, rebuild")
        sys.exit(1)

    command = sys.argv[1]
    tracker_path = os.path.expanduser(sys.argv[2])

    if command == "dedup-set":
        ids = get_dedup_set(tracker_path)
        print(json.dumps(ids))

    elif command == "append":
        if len(sys.argv) < 4:
            print("Usage: python3 tracker_utils.py append <tracker_path> <new_rows.json>")
            sys.exit(1)
        result = append_rows(tracker_path, sys.argv[3])
        print(json.dumps(result, indent=2))

    elif command == "rebuild":
        result = rebuild(tracker_path)
        print(json.dumps(result, indent=2))

    else:
        print(f"Unknown command: {command}")
        sys.exit(1)
